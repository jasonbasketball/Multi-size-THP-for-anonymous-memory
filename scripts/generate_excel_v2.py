#!/usr/bin/env python3
"""
Generate Excel file with Multi-size THP integration information - V2
"""

import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_report():
    """Create comprehensive Excel report for Multi-size THP integration"""
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: 项目概述 (Project Overview)
    ws1 = wb.active
    ws1.title = "项目概述"
    
    ws1['A1'] = "Multi-size THP for Anonymous Memory - 合入信息"
    ws1['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws1.merge_cells('A1:D1')
    
    overview_data = [
        (3, 1, "项目名称", Font(bold=True)),
        (3, 2, "Multi-size THP for Anonymous Memory", None),
        (4, 1, "功能描述", Font(bold=True)),
        (4, 2, "多尺寸透明大页为匿名内存提供运行时选择功能", None),
        (5, 1, "源内核版本", Font(bold=True)),
        (5, 2, "Linux 6.8", None),
        (6, 1, "目标内核版本", Font(bold=True)),
        (6, 2, "OpenEuler 6.6", None),
        (7, 1, "主要贡献者", Font(bold=True)),
        (7, 2, "Ryan Roberts (ARM)", None),
        (8, 1, "特性状态", Font(bold=True)),
        (8, 2, "已合入主线 (Mainline Merged)", None),
        (9, 1, "生成日期", Font(bold=True)),
        (9, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None),
    ]
    
    for row, col, value, font in overview_data:
        cell = ws1.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Features section
    ws1.cell(row=11, column=1, value="功能特点").font = Font(bold=True, size=12)
    features = [
        "支持多种大小的透明大页 (64KB, 128KB, 256KB, 512KB, 1MB, 2MB)",
        "通过sysfs接口进行运行时配置",
        "提供每种大小的分配统计信息",
        "支持ARM64架构的连续页表项优化",
        "改进页错误处理性能",
        "支持NUMA平衡",
    ]
    
    for idx, feature in enumerate(features, start=12):
        ws1.cell(row=idx, column=1, value=str(idx-11))
        ws1.cell(row=idx, column=2, value=feature)
    
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 50
    
    # Sheet 2: 提交列表 (Commit List)
    ws2 = wb.create_sheet("提交列表")
    
    headers = ["序号", "提交标题", "功能描述", "子系统", "影响文件数", "预估代码行数"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    commit_details = [
        ("mm: add opt-in for multi-size THP for anon_fault", "引入对匿名内存多尺寸THP的支持", "mm", 3, "~500"),
        ("mm: thp: Introduce multi-size THP sysfs interface", "添加sysfs接口以控制多尺寸THP", "mm", 2, "~300"),
        ("mm: allow deferred splitting of arbitrary anon large folios", "允许延迟拆分任意大小的匿名大页folios", "mm", 2, "~200"),
        ("mm: add per-order mTHP alloc_success and alloc_fail counters", "添加每种大小的mTHP分配计数器", "mm", 2, "~150"),
        ("mm: implement split folios for multi-size THP", "实现多尺寸THP的folio拆分功能", "mm", 1, "~250"),
        ("arm64/mm: Wire up PTE_CONT for user mappings", "为ARM64连接PTE_CONT支持", "arm64", 2, "~400"),
        ("mm: add new KSM process and sysfs knobs", "添加新的KSM进程和sysfs配置", "mm", 2, "~200"),
        ("mm: support multi-size THP numa balancing", "支持多尺寸THP的NUMA平衡", "mm", 2, "~180"),
        ("mm: add large folio shmem swapin support", "添加大folio共享内存的换入支持", "mm", 2, "~220"),
        ("mm: allow THP to be configured by sysfs", "允许通过sysfs配置THP参数", "mm", 2, "~150"),
    ]
    
    for idx, (title, desc, subsys, files, lines) in enumerate(commit_details, start=2):
        ws2.cell(row=idx, column=1, value=idx-1)
        ws2.cell(row=idx, column=2, value=title)
        ws2.cell(row=idx, column=3, value=desc)
        ws2.cell(row=idx, column=4, value=subsys)
        ws2.cell(row=idx, column=5, value=files)
        ws2.cell(row=idx, column=6, value=lines)
        
        for col in range(1, 7):
            cell = ws2.cell(row=idx, column=col)
            cell.border = border
            if col in [1, 5]:
                cell.alignment = Alignment(horizontal='center')
    
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 45
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 15
    
    # Sheet 3: 代码统计 (Code Statistics)
    ws3 = wb.create_sheet("代码统计")
    
    ws3['A1'] = "代码变更统计"
    ws3['A1'].font = Font(bold=True, size=14)
    
    ws3.cell(row=3, column=1, value="统计项").font = subheader_font
    ws3.cell(row=3, column=1).fill = subheader_fill
    ws3.cell(row=3, column=2, value="数量").font = subheader_font
    ws3.cell(row=3, column=2).fill = subheader_fill
    ws3.cell(row=3, column=3, value="说明").font = subheader_font
    ws3.cell(row=3, column=3).fill = subheader_fill
    
    stats_data = [
        ("提交总数", "10", "主要提交数量"),
        ("修改文件总数", "~15-20", "包括源文件、头文件和文档"),
        ("新增代码行数", "~2,500", "新增功能代码"),
        ("删除代码行数", "~300", "移除或重构的代码"),
        ("净增代码行数", "~2,200", "实际增加的代码量"),
    ]
    
    for idx, (item, value, note) in enumerate(stats_data, start=4):
        ws3.cell(row=idx, column=1, value=item).font = Font(bold=True)
        ws3.cell(row=idx, column=2, value=value)
        ws3.cell(row=idx, column=3, value=note)
    
    ws3.cell(row=10, column=1, value="主要修改文件").font = Font(bold=True, size=12)
    
    files = [
        ("mm/huge_memory.c", "核心文件", "主要THP实现"),
        ("mm/memory.c", "核心文件", "页错误处理"),
        ("mm/rmap.c", "核心文件", "反向映射处理"),
        ("include/linux/huge_mm.h", "头文件", "接口定义"),
        ("arch/arm64/mm/mmu.c", "架构文件", "ARM64特定支持"),
        ("Documentation/admin-guide/mm/transhuge.rst", "文档", "用户文档"),
    ]
    
    for idx, (path, ftype, desc) in enumerate(files, start=11):
        ws3.cell(row=idx, column=1, value=path)
        ws3.cell(row=idx, column=2, value=ftype)
        ws3.cell(row=idx, column=3, value=desc)
    
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 30
    
    # Sheet 4: 文件清单 (File List)
    ws4 = wb.create_sheet("文件清单")
    
    headers = ["序号", "文件路径", "文件类型", "主要改动", "改动规模"]
    for col, header in enumerate(headers, start=1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    file_list = [
        ("mm/huge_memory.c", "源代码", "多尺寸THP核心实现、sysfs接口、统计", "大"),
        ("mm/memory.c", "源代码", "页错误路径支持多尺寸THP", "中"),
        ("mm/rmap.c", "源代码", "延迟拆分支持", "小"),
        ("mm/mprotect.c", "源代码", "NUMA平衡支持", "小"),
        ("mm/shmem.c", "源代码", "大folio换入支持", "中"),
        ("mm/ksm.c", "源代码", "KSM更新", "小"),
        ("include/linux/huge_mm.h", "头文件", "接口和数据结构定义", "中"),
        ("include/linux/ksm.h", "头文件", "KSM接口更新", "小"),
        ("arch/arm64/mm/mmu.c", "源代码", "ARM64 PTE_CONT支持", "中"),
        ("arch/arm64/include/asm/pgtable.h", "头文件", "ARM64页表定义", "小"),
        ("Documentation/admin-guide/mm/transhuge.rst", "文档", "mTHP用户文档", "中"),
    ]
    
    for idx, (path, ftype, changes, scale) in enumerate(file_list, start=2):
        ws4.cell(row=idx, column=1, value=idx-1)
        ws4.cell(row=idx, column=2, value=path)
        ws4.cell(row=idx, column=3, value=ftype)
        ws4.cell(row=idx, column=4, value=changes)
        ws4.cell(row=idx, column=5, value=scale)
        
        for col in range(1, 6):
            cell = ws4.cell(row=idx, column=col)
            cell.border = border
            if col in [1, 5]:
                cell.alignment = Alignment(horizontal='center')
    
    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 40
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 40
    ws4.column_dimensions['E'].width = 12
    
    # Sheet 5: 合入计划 (Integration Plan)
    ws5 = wb.create_sheet("合入计划")
    
    ws5['A1'] = "OpenEuler 6.6 合入计划"
    ws5['A1'].font = Font(bold=True, size=14)
    
    ws5.cell(row=3, column=1, value="阶段").font = subheader_font
    ws5.cell(row=3, column=1).fill = subheader_fill
    ws5.cell(row=3, column=2, value="任务").font = subheader_font
    ws5.cell(row=3, column=2).fill = subheader_fill
    ws5.cell(row=3, column=3, value="注意事项").font = subheader_font
    ws5.cell(row=3, column=3).fill = subheader_fill
    ws5.cell(row=3, column=4, value="状态").font = subheader_font
    ws5.cell(row=3, column=4).fill = subheader_fill
    
    plan_data = [
        ("1. 准备阶段", "分析OpenEuler 6.6内核版本", "确认folio基础设施是否存在", "待完成"),
        ("", "检查依赖的内核功能", "验证必要的mm子系统功能", "待完成"),
        ("", "准备测试环境", "搭建编译和测试环境", "待完成"),
        ("2. 移植阶段", "应用核心mm补丁", "可能需要适配API差异", "待完成"),
        ("", "应用ARM64架构补丁", "检查ARM64特定代码兼容性", "待完成"),
        ("", "应用文档更新", "翻译和本地化文档", "待完成"),
        ("3. 测试阶段", "编译测试", "解决编译错误", "待完成"),
        ("", "功能测试", "验证mTHP基本功能", "待完成"),
        ("", "性能测试", "对比性能提升", "待完成"),
        ("", "稳定性测试", "长时间压力测试", "待完成"),
        ("4. 优化阶段", "性能优化", "针对OpenEuler优化", "待完成"),
        ("", "代码审查", "符合OpenEuler编码规范", "待完成"),
        ("5. 发布阶段", "文档整理", "更新发行说明", "待完成"),
        ("", "补丁提交", "提交到OpenEuler仓库", "待完成"),
    ]
    
    for idx, (stage, task, note, status) in enumerate(plan_data, start=4):
        ws5.cell(row=idx, column=1, value=stage)
        ws5.cell(row=idx, column=2, value=task)
        ws5.cell(row=idx, column=3, value=note)
        ws5.cell(row=idx, column=4, value=status)
    
    ws5.column_dimensions['A'].width = 15
    ws5.column_dimensions['B'].width = 30
    ws5.column_dimensions['C'].width = 35
    ws5.column_dimensions['D'].width = 12
    
    # Sheet 6: 风险评估 (Risk Assessment)
    ws6 = wb.create_sheet("风险评估")
    
    headers = ["风险项", "风险等级", "影响范围", "缓解措施"]
    for col, header in enumerate(headers, start=1):
        cell = ws6.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    risks = [
        ("API兼容性问题", "中", "mm子系统", "仔细检查API变化，必要时适配"),
        ("folio基础设施缺失", "高", "核心功能", "评估是否需要先移植folio支持"),
        ("性能回归", "中", "整体性能", "充分的性能测试和基准对比"),
        ("内存管理稳定性", "中", "系统稳定性", "长时间压力测试"),
        ("ARM64特定功能", "低", "ARM64平台", "在ARM64平台充分测试"),
        ("与OpenEuler补丁冲突", "中", "补丁合并", "仔细解决冲突，保留OpenEuler特性"),
    ]
    
    for idx, (risk, level, impact, mitigation) in enumerate(risks, start=2):
        ws6.cell(row=idx, column=1, value=risk)
        ws6.cell(row=idx, column=2, value=level)
        ws6.cell(row=idx, column=3, value=impact)
        ws6.cell(row=idx, column=4, value=mitigation)
        
        for col in range(1, 5):
            cell = ws6.cell(row=idx, column=col)
            cell.border = border
            if col == 2:
                cell.alignment = Alignment(horizontal='center')
                if level == "高":
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif level == "中":
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    
    ws6.column_dimensions['A'].width = 25
    ws6.column_dimensions['B'].width = 12
    ws6.column_dimensions['C'].width = 20
    ws6.column_dimensions['D'].width = 40
    
    # Save the workbook
    filename = "Multi-size_THP_Integration_Report.xlsx"
    wb.save(filename)
    print(f"✓ Excel报告已生成: {filename}")
    
    return filename

if __name__ == "__main__":
    print("正在生成Excel集成报告...\n")
    filename = create_excel_report()
    print(f"\n报告包含以下工作表:")
    print("  1. 项目概述 - 功能特点和基本信息")
    print("  2. 提交列表 - 所有相关提交的详细信息")
    print("  3. 代码统计 - 代码变更量统计")
    print("  4. 文件清单 - 修改文件的详细列表")
    print("  5. 合入计划 - OpenEuler 6.6集成计划")
    print("  6. 风险评估 - 潜在风险和缓解措施")
    print(f"\n✓ 完成！")

